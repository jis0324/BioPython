#This line imports the SeqIO class from the biopython module which helps us to read the sequences from the fasta file.
from Bio import SeqIO
#This line imports the openpyxl module which is needed for working with Excel files.
import openpyxl 

#This function takes a number between 0 and 1 as the decimal pecent and shows the progress bar for that percent
def progresBar(percent):
    print(str(int(percent * 100)).zfill(3) + ": [", end='')
    print(int(percent * 100) * "*" + "]")

#This line reads all of the records from the "Example.fasta" file.
records = list(SeqIO.parse(r"input.fasta", "fasta"))

#This line defines two variables used for showing the progress bar.
counter = 0; lastP = -1
#"SNPs" variable is a dictionary which consists of the information of the SNPs.  The structure of this dictionary is as follows:
#SNPs = {1: [[2, 'G', 'T', 2103], [5, 'A', 'T', 1027]], 3: [[1, 'T', 'A', 4510], [8, 'A', 'T', 2222]], ...}
#The keys are the number of the reference sequence and the values are lists of lists which contains 4 elements: 1. The compared sequence, 2. The character
#in the reference sequence, 3. the character in the compared sequence, and 4. the position of the character in the sequence.
SNPs = {}

#The following loop is using to compare the sequences read from the file with the reference sequence to find SNPs
#It compares all sequences with the reference sequence and then loops over the other ones and compares them.
#Each sheet of the Excel file is related to one of the sequence and its name is the number of the reference sequence in the fasta file.
refSeq = records[0].seq
for i in range(1, len(records)):
    #This line adds a key for the current reference sequence to "SNPs"
    SNPs[i] = []
    #Now that we have defined the reference sequence, we should define the next sequence to be compared with the reference one
    seq1 = records[i].seq
    #"mini" is the minimum length of these two sequences (seq1 and seq2)
    mini = min(len(refSeq), len(seq1))

    # position number without gaps.
    ref_position = 0
    seq_position = 0
    #Now it's time to loop over each character of the sequences and compare them to each other
    for k in range(mini):
        # increase position if refSeq's kth character is not gap
        if refSeq[k] != "-":
            ref_position += 1
        # increase position if seq's kth character is not gap
        if seq1[k] != "-":
            seq_position += 1
        #This "if" statement determines whether the k-th character of the sequences are the same or not. If the answer is NO, then it's a SNP and
        #we have to store that in the "SNPs" variable based on the above mentioned format.
        if refSeq[k] != seq1[k]:
            # if kth character of reference or rest sequence is gap, ignore.
            if refSeq[k] == "-" or records[i][k] == "-":
                continue
            else:
                SNPs[i].append([refSeq[k], seq1[k], ref_position, seq_position])

    #"counter" stores the number of the sequences compared to each other so that we can compute the percentage of the work done.
    counter = counter + 1
    #"percent" is the "counter" variable divided by the number of the total comparisons which is the square of the number of the total records
    percent = counter / len(records)
    #If the percentage has changed, it shows a progress in the progres bar.
    if int(percent * 100) > lastP:
        lastP = int(percent * 100)
        progresBar(percent)
         
            
#"mutations" is a dictionary which stores the SNPs mutations with the following format:
#mutations = {1:['GT': 0.45, 'AT': 0.5, 'AC': 0.88], 2:['AT': 0.2], 3:['CT': 0.32, 'TA': 0.4]}
#The keys in this dictionary are the position in the sequences and the values are a list of SNP mutations in that position in all of the sequences.
#The number in front of each SNPs indicates the percentage of that mutation.
mutations = {}
RegSNPs = {}
# position number without gaps.
position = 0
for k in range(len(refSeq)):
    #First, we identify the k-th character in the reference sequence (refChar)
    refChar = refSeq[k]
    # increase position if refChar is not gap
    if refChar == "-":
        continue
    else:
        position += 1
    #Then, we search for the SNPs at k-th character and store their number of occurance in "kSNP" dictionary.
    kSNP = {}
    kSNPPosition = {}
    for i in range(1, len(records)):
        if k < len(records[i].seq):
            #If the k-th character in the compared sequence is different form the "refChar", we have to consider it in the "kSNP" dictionary
            if records[i].seq[k] != refChar and records[i].seq[k] != "-":
                #If this SNP was already in the kSNP, we only add 1 to its number of occurance.
                if refChar + records[i].seq[k] in kSNP.keys():
                    kSNP[refChar + records[i].seq[k]] = kSNP[refChar + records[i].seq[k]] + 1
                    kSNPPosition[refChar + records[i].seq[k]].append(i)
                #But if it is their first time of occurance, we add this SNP to the "kSNP" keys and assign 1 to its value.
                else:
                    kSNP[refChar + records[i].seq[k]] = 1
                    kSNPPosition[refChar + records[i].seq[k]] = [i]
    #After finishin of this procedure, we search for that SNPs with occurance less than 1% in the k-th position and store the SNP and their percentage
    #in the "mutation[k]".
    mutations[k] = []
    RegSNPs[k] = []
    for key in kSNP.keys():
        # print(kSNP[key] / len(records) * 100)
        if kSNP[key] / len(records) * 100 < 1:
            mutations[k].append([key, round(kSNP[key] / len(records) * 100, 2), kSNPPosition[key], position])
        else:
            RegSNPs[k].append([key, round(kSNP[key] / len(records) * 100, 2), kSNPPosition[key], position])

#The next two lines, deletes all the positions with no mutation from "mutations" dictionary
tmp = {k: v for k, v in list(mutations.items()) if len(v) != 0}
mutations = tmp

tmp = {k: v for k, v in list(RegSNPs.items()) if len(v) != 0}
RegSNPs = tmp

#This line opens an Excel workbook and stores that in "wbm" variable
wbm = openpyxl.Workbook()

#First, we add a sheet named "All-Mutations" and store all the mutations there.
wbm.create_sheet(index = 1, title = "All-Mutations")
sheet = wbm["All-Mutations"]
sheet.cell(1, 1).value = "Position"
sheet.cell(1, 2).value = "SNP"
sheet.cell(1, 3).value = "Percentage"
row = 2
for pos in mutations.keys():
    for item in mutations[pos]:
        sheet.cell(row, 1).value = str(pos)
        sheet.cell(row, 2).value = str(item[0])
        sheet.cell(row, 3).value = str(item[1])
        row = row + 1

for i in range(1, len(records)):
    #This line creates a sheet for each sequence
    wbm.create_sheet(index = i+1, title = str(i+1))
    #This line puts the created sheet in the "sheet" variable
    sheet = wbm[str(i+1)]
    #The next lines adds the title of each column to the Excel file.
    #The first column is the position of the mutation
    sheet.cell(1, 1).value = "Position"
    #The second column is the SNP pair
    sheet.cell(1, 2).value = "SNP"
    #The third column is the percentage of the SNP
    sheet.cell(1, 3).value = "Percentage"

#Now that each sequence has a sheet, we add the mutations in that sequence to their sheet.
for pos in mutations.keys():
    for item in mutations[pos]:
        for rec in item[2]:
            sheet = wbm[str(int(rec)+1)]
            row = len(sheet['A']) + 1
            sheet.cell(row, 1).value = str(pos)
            sheet.cell(row, 2).value = str(item[0])
            sheet.cell(row, 3).value = str(item[1])
            
#Finally we save the Excel workbook with the name of "Mutations.xlsx"
wbm.save(r"C:\Users\a-s-j\Desktop\Experemint\scripts\Mutations.xlsx")

#*********************************************************************
#*********************************************************************
#This line opens an Excel workbook and stores that in "wbm" variable
wbm = openpyxl.Workbook()

#First, we add a sheet named "All-Mutations" and store all the mutations there.
wbm.create_sheet(index = 1, title = "All-RegularSNPs")
sheet = wbm["All-RegularSNPs"]
sheet.cell(1, 1).value = "Position"
sheet.cell(1, 2).value = "SNP"
sheet.cell(1, 3).value = "Percentage"
row = 2
for pos in RegSNPs.keys():
    for item in RegSNPs[pos]:
        sheet.cell(row, 1).value = str(pos)
        sheet.cell(row, 2).value = str(item[0])
        sheet.cell(row, 3).value = str(item[1])
        row = row + 1

for i in range(1, len(records)):
    #This line creates a sheet for each sequence
    wbm.create_sheet(index = i+1, title = str(i+1))
    #This line puts the created sheet in the "sheet" variable
    sheet = wbm[str(i+1)]
    #The next lines adds the title of each column to the Excel file.
    #The first column is the position of the mutation
    sheet.cell(1, 1).value = "Position"
    #The second column is the SNP pair
    sheet.cell(1, 2).value = "SNP"
    #The third column is the percentage of the SNP
    sheet.cell(1, 3).value = "Percentage"

#Now that each sequence has a sheet, we add the RegSNPs in that sequence to their sheet.
for pos in RegSNPs.keys():
    for item in RegSNPs[pos]:
        for rec in item[2]:
            sheet = wbm[str(int(rec)+1)]
            row = len(sheet['A']) + 1
            sheet.cell(row, 1).value = str(pos)
            sheet.cell(row, 2).value = str(item[0])
            sheet.cell(row, 3).value = str(item[1])
            
#Finally we save the Excel workbook with the name of "Mutations.xlsx"
wbm.save(r"C:\Users\a-s-j\Desktop\Experemint\scripts\RegularSNPs.xlsx")


#*********************************************************************
#*********************************************************************

#This line opens an Excel workbook and stores that in "wb" variable
wb = openpyxl.Workbook()

wb.create_sheet(index = 1, title = "All-SNPs")
sheet = wb["All-SNPs"]
sheet.cell(1, 1).value = "Sequence"
sheet.cell(1, 2).value = "k-th character in reference sequence"
sheet.cell(1, 3).value = "k-th character in the compared sequence"
sheet.cell(1, 4).value = "k"
row = 2
for key1 in SNPs.keys():
    for item in SNPs[key1]:
        sheet.cell(row, 1).value = str(key1)
        sheet.cell(row, 2).value = str(item[0])
        sheet.cell(row, 3).value = str(item[1])
        sheet.cell(row, 4).value = str(item[2])
        row = row + 1

# create new sheet named as "relations"
wb.create_sheet(index = 2, title = "relations")
# active created sheet - relations.
sheet = wb["relations"]
# write header on sheet
sheet.cell(1, 1).value = "seqId"
sheet.cell(1, 2).value = "position of snps in sequence"
sheet.cell(1, 3).value = "position of snps in refernce"
sheet.cell(1, 4).value = "position and charecters in reference"
sheet.cell(1, 5).value = "position of reference and character in sequence"
sheet.cell(1, 6).value = "mapping (position of reference, character of reference, character of sequence)"
sheet.cell(1, 7).value = "num of children"
row = 2

# Iterate SNPs
for key2, value2 in SNPs.items():
    # `seqId` column value. here key2 is seqId
    sheet.cell(row, 1).value = str(key2)

    # position list of snps in sequence
    """
    [3, 4, 5, 7, 8, 11, 12, 14, 16, 18, 19, 20, 22, 23, 249, 3045, 6320, 11091, 13738, 14416, 18885, 23411, 23937, 25571, 25623, 26743, 28319]
    """
    seq_snps_pos = list()

    # position list of snps in refernce
    """
    [1, 2, 3, 5, 6, 9, 10, 12, 14, 16, 17, 18, 20, 21, 247, 3043, 6318, 11089, 13736, 14414, 18883, 23409, 23935, 25569, 25621, 26741, 28317]
    """
    ref_snps_pos = list()

    # list of position and charecters in reference sequence
    """
    {1: 'C', 2: 'C', 3: 'T', 5: 'G', 6: 'G', 9: 'G', 10: 'G', 12: 'A', 14: 'A', 16: 'A', 17: 'C', 18: 'C', 20: 'T', 21: 'A', 247: 'C', 3043: 'C', 6318: 'A', 11089: 'T', 13736: 'T', 14414: 'C', 18883: 'C', 23409: 'A', 23935: 'T', 25569: 'G', 25621: 'G', 26741: 'C', 28317: 'T'}
    """
    ref_snps_pos_char = dict()

    # mapping (position of reference, character of reference, character of sequence)
    """
    {1: 'A', 2: 'A', 3: 'G', 5: 'C', 6: 'C', 9: 'A', 10: 'C', 12: 'G', 14: 'C', 16: 'C', 17: 'T', 18: 'A', 20: 'A', 21: 'C', 247: 'T', 3043: 'T', 6318: 'C', 11089: 'G', 13736: 'C', 14414: 'T', 18883: 'T', 23409: 'G', 23935: 'C', 25569: 'T', 25621: 'A', 26741: 'T', 28317: 'C'}
    """
    ref_snps_pos_seq_snps_char = dict()

    # mapping (position of reference, character of reference, character of sequence)
    """
    {1: 'C->A', 2: 'C->A', 3: 'T->G', 5: 'G->C', 6: 'G->C', 9: 'G->A', 10: 'G->C', 12: 'A->G', 14: 'A->C', 16: 'A->C', 17: 'C->T', 18: 'C->A', 20: 'T->A', 21: 'A->C', 247: 'C->T', 3043: 'C->T', 6318: 'A->C', 11089: 'T->G', 13736: 'T->C', 14414: 'C->T', 18883: 'C->T', 23409: 'A->G', 23935: 'T->C', 25569: 'G->T', 25621: 'G->A', 26741: 'C->T', 28317: 'T->C'}
    """
    mapping = dict()
    
    for item in value2:
        # append the position of snps in sequence
        seq_snps_pos.append(item[3])
        # append the position of snps in refernce
        ref_snps_pos.append(item[2])
        # position and charecters in reference
        ref_snps_pos_char[item[2]] = item[0]
        # position of reference and character in sequence
        ref_snps_pos_seq_snps_char[item[2]] = item[1]
        # mapping (position of reference, character of reference, character of sequence)
        mapping[item[2]] = item[0]+"->"+item[1]

    # write `position of snps in sequence` column
    sheet.cell(row, 2).value = str(seq_snps_pos)
    # write `position of snps in reference` column
    sheet.cell(row, 3).value = str(ref_snps_pos)
    # write `position and charecters in reference` column
    sheet.cell(row, 4).value = str(ref_snps_pos_char)
    # write `position of reference and character in sequence` column
    sheet.cell(row, 5).value = str(ref_snps_pos_seq_snps_char)
    # write `mapping (position of reference, character of reference, character of sequence)` column
    sheet.cell(row, 6).value = str(mapping)

    # child count
    count = 0
    for key3, value3 in SNPs.items():
        # ignore between same elements
        if key2 == key3:
            continue
        
        value3 = [item[:3] for item in value3]
        # if value3 contains value2,  increase children count
        if all([item[:3] in value3 for item in value2]):
            count += 1

    # write children column
    sheet.cell(row, 7).value = str(count)

    # increase row
    row = row + 1

for key in SNPs.keys():
    #This line creates a sheet with the name of the reference sequence
    wb.create_sheet(index = int(key)+2 , title = str(int(key)+1))
    #This line puts the created sheet in the "sheet" variable
    sheet = wb[str(int(key)+1)]
    #The next lines adds the title of each column to the Excel file.
    #The second column is the k-th character of the reference sequence
    sheet.cell(1, 1).value = "k-th item in the reference sequence"
    #The third column is the k-th character of the compared sequence
    sheet.cell(1, 2).value = "k-th item in the sequence"
    #And the fourth column is the position of the SNP in the sequence (k)
    sheet.cell(1, 3).value = "k"
    #"row" variable shows the current last row to put the new SNP there. So at the beginning, it is equal to 1 and its value is updated throghout the loop
    row = 2
    #The next loop adds the information of this SNP in to the Excel file based on the above mentioned definition. 
    for item in SNPs[key]:
        sheet.cell(row, 1).value = item[0]
        sheet.cell(row, 2).value = item[1]
        sheet.cell(row, 3).value = item[2]
        #Then we updated the "row" variable so that we are at the beginning of the next empty row to add information
        row = row + 1

#Finally we save the Excel workbook with the name of "Output.xlsx"
wb.save(r"C:\Users\a-s-j\Desktop\Experemint\scripts\Output.xlsx")
