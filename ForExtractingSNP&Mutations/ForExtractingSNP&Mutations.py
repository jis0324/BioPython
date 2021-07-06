#This line imports the SeqIO class from the biopython module which helps us to read the sequences from the fasta file.
from Bio import SeqIO
#This line imports the openpyxl module which is needed for working with Excel files.
import openpyxl 

#This function takes a number between 0 and 1 as the decimal pecent and shows the progress bar for that percent
def progresBar(percent):
    print(str(int(percent * 100)).zfill(3) + ": [", end='')
    print(int(percent * 100) * "*" + "]")

#This line reads all of the records from the "Example.fasta" file.
records = list(SeqIO.parse("Example.fasta", "fasta"))

#This line defines two variables used for showing the progress bar.
counter = 0; lastP = -1
#"SNPs" variable is a dictionary which consists of the information of the SNPs.  The structure of this dictionary is as follows:
#SNPs = {1: [[2, 'G', 'T', 2103], [5, 'A', 'T', 1027]], 3: [[1, 'T', 'A', 4510], [8, 'A', 'T', 2222]], ...}
#The keys are the number of the reference sequence and the values are lists of lists which contains 4 elements: 1. The compared sequence, 2. The character
#in the reference sequence, 3. the character in the compared sequence, and 4. the position of the character in the sequence.
SNPs = {}

#The following loop is using to compare the sequences read from the file with the reference sequence to find SNPs
#It compares all sequences with the reference sequence and then loops over the other ones and compares them.
#Each sheet of the Excel file is related to one of the sequence and its name is the number of the reference sequence in the fasta file.
refSeq = records[0].seq
for i in range(1, len(records)):
    #This line adds a key for the current reference sequence to "SNPs"
    SNPs[i] = []
    #Now that we have defined the reference sequence, we should define the next sequence to be compared with the reference one
    seq1 = records[i].seq
    #"mini" is the minimum length of these two sequences (seq1 and seq2)
    mini = min(len(refSeq), len(seq1))

    # position number without gaps.
    ref_position = 0
    seq_position = 0
    #Now it's time to loop over each character of the sequences and compare them to each other
    for k in range(mini):
        # increase position if refSeq's kth character is not gap
        if refSeq[k] != "-":
            ref_position += 1
        # increase position if seq's kth character is not gap
        if seq1[k] != "-":
            seq_position += 1
        #This "if" statement determines whether the k-th character of the sequences are the same or not. If the answer is NO, then it's a SNP and
        #we have to store that in the "SNPs" variable based on the above mentioned format.
        if refSeq[k] != seq1[k]:
            # if kth character of reference or rest sequence is gap, ignore.
            if refSeq[k] == "-" or records[i][k] == "-":
                continue
            else:
                SNPs[i].append([refSeq[k], seq1[k], ref_position, seq_position])

    #"counter" stores the number of the sequences compared to each other so that we can compute the percentage of the work done.
    counter = counter + 1
    #"percent" is the "counter" variable divided by the number of the total comparisons which is the square of the number of the total records
    percent = counter / len(records)
    #If the percentage has changed, it shows a progress in the progres bar.
    if int(percent * 100) > lastP:
        lastP = int(percent * 100)
        progresBar(percent)
         
            
#"mutations" is a dictionary which stores the SNPs mutations with the following format:
#mutations = {1:['GT': 0.45, 'AT': 0.5, 'AC': 0.88], 2:['AT': 0.2], 3:['CT': 0.32, 'TA': 0.4]}
#The keys in this dictionary are the position in the sequences and the values are a list of SNP mutations in that position in all of the sequences.
#The number in front of each SNPs indicates the percentage of that mutation.
mutations = {}
# position number without gaps.
position = 0
for k in range(len(refSeq)):
    #First, we identify the k-th character in the reference sequence (refChar)
    refChar = refSeq[k]
    # increase position if refChar is not gap
    if refChar == "-":
        continue
    else:
        position += 1
    #Then, we search for the SNPs at k-th character and store their number of occurance in "kSNP" dictionary.
    kSNP = {}
    kSNPPosition = {}
    for i in range(1, len(records)):
        if k < len(records[i].seq):
            # check gap in sequence
            if records[i].seq[k] == "-":
                continue

            #If the k-th character in the compared sequence is different form the "refChar", we have to consider it in the "kSNP" dictionary
            if records[i].seq[k] != refChar:
                #If this SNP was already in the kSNP, we only add 1 to its number of occurance.
                if refChar + records[i].seq[k] in kSNP.keys():
                    kSNP[refChar + records[i].seq[k]] = kSNP[refChar + records[i].seq[k]] + 1
                    kSNPPosition[refChar + records[i].seq[k]].append(i)
                #But if it is their first time of occurance, we add this SNP to the "kSNP" keys and assign 1 to its value.
                else:
                    kSNP[refChar + records[i].seq[k]] = 1
                    kSNPPosition[refChar + records[i].seq[k]] = [i]
    #After finishin of this procedure, we search for that SNPs with occurance less than 1% in the k-th position and store the SNP and their percentage
    #in the "mutation[k]".
    mutations[k] = []
    for key in kSNP.keys():
        if kSNP[key] / len(records) * 100 <= 1:
            mutations[k].append([key, round(kSNP[key] / len(records) * 100, 2), kSNPPosition[key]], position)

#The next two lines, deletes all the positions with no mutation from "mutations" dictionary
tmp = {k: v for k, v in list(mutations.items()) if len(v) != 0}
mutations = tmp

#This line opens an Excel workbook and stores that in "wbm" variable
wbm = openpyxl.Workbook()

#First, we add a sheet named "All-Mutations" and store all the mutations there.
wbm.create_sheet(index = 1, title = "All-Mutations")
sheet = wbm["All-Mutations"]
sheet.cell(1, 1).value = "Position"
sheet.cell(1, 2).value = "SNP"
sheet.cell(1, 3).value = "Percentage"
row = 2
for pos in mutations.keys():
    for item in mutations[pos]:
        sheet.cell(row, 1).value = str(pos)
        sheet.cell(row, 2).value = str(item[0])
        sheet.cell(row, 3).value = str(item[1])
        row = row + 1

for i in range(1, len(records)):
    #This line creates a sheet for each sequence
    wbm.create_sheet(index = i+1, title = str(i+1))
    #This line puts the created sheet in the "sheet" variable
    sheet = wbm[str(i+1)]
    #The next lines adds the title of each column to the Excel file.
    #The first column is the position of the mutation
    sheet.cell(1, 1).value = "Position"
    #The second column is the SNP pair
    sheet.cell(1, 2).value = "SNP"
    #The third column is the percentage of the SNP
    sheet.cell(1, 3).value = "Percentage"

#Now that each sequence has a sheet, we add the mutations in that sequence to their sheet.
for pos in mutations.keys():
    for item in mutations[pos]:
        for rec in item[2]:
            sheet = wbm[str(int(rec)+1)]
            row = len(sheet['A']) + 1
            sheet.cell(row, 1).value = str(pos)
            sheet.cell(row, 2).value = str(item[0])
            sheet.cell(row, 3).value = str(item[1])
            
#Finally we save the Excel workbook with the name of "Mutations.xlsx"
wbm.save("Mutations.xlsx")

#This line opens an Excel workbook and stores that in "wb" variable
wb = openpyxl.Workbook()

wb.create_sheet(index = 1, title = "All-SNPs")
sheet = wb["All-SNPs"]
sheet.cell(1, 1).value = "Sequence"
sheet.cell(1, 2).value = "k-th character in reference sequence"
sheet.cell(1, 3).value = "k-th character in the compared sequence"
sheet.cell(1, 4).value = "k"
row = 2
for key1 in SNPs.keys():
    for item in SNPs[key1]:
        sheet.cell(row, 1).value = str(key1)
        sheet.cell(row, 2).value = str(item[0])
        sheet.cell(row, 3).value = str(item[1])
        sheet.cell(row, 4).value = str(item[2])
        row = row + 1

wb.create_sheet(index = 2, title = "relations")
sheet = wb["relations"]
sheet.cell(1, 1).value = "seqId"
sheet.cell(1, 2).value = "position of snps in sequence"
sheet.cell(1, 3).value = "position of snps in refernce"
sheet.cell(1, 4).value = "position and charecters in reference"
sheet.cell(1, 5).value = "position of reference and character in sequence"
sheet.cell(1, 6).value = "mapping (position of reference, character of reference, character of sequence)"
sheet.cell(1, 7).value = "num of children"
row = 2
for key2, value2 in SNPs.items():
    sheet.cell(row, 1).value = str(key2)

    seq_snps_pos = list()
    ref_snps_pos = list()
    ref_snps_pos_char = dict()
    ref_snps_pos_seq_snps_char = dict()
    mapping = dict()
    for item in value2:
        # position of snps in sequence
        seq_snps_pos.append(item[3])
        # position of snps in refernce
        ref_snps_pos.append(item[2])
        # position and charecters in reference
        ref_snps_pos_char[item[2]] = item[0]
        # position of reference and character in sequence
        ref_snps_pos_seq_snps_char[item[2]] = item[1]
        # mapping (position of reference, character of reference, character of sequence)
        mapping[item[2]] = item[0]+"->"+item[1]

    sheet.cell(row, 2).value = str(seq_snps_pos)
    sheet.cell(row, 3).value = str(ref_snps_pos)
    sheet.cell(row, 4).value = str(ref_snps_pos_char)
    sheet.cell(row, 5).value = str(ref_snps_pos_seq_snps_char)
    sheet.cell(row, 6).value = str(mapping)

    count = 0
    for key3, value3 in SNPs.items():
        # ignore between same elements
        if key2 == key3:
            continue
        
        value3 = [item[:3] for item in value3]
        # if value3 contains value2,  increase children count
        if all([item[:3] in value3 for item in value2]):
            count += 1

    sheet.cell(row, 7).value = str(count)

    row = row + 1

for key in SNPs.keys():
    #This line creates a sheet with the name of the reference sequence
    wb.create_sheet(index = int(key)+2 , title = str(int(key)+1))
    #This line puts the created sheet in the "sheet" variable
    sheet = wb[str(int(key)+1)]
    #The next lines adds the title of each column to the Excel file.
    #The second column is the k-th character of the reference sequence
    sheet.cell(1, 1).value = "k-th item in the reference sequence"
    #The third column is the k-th character of the compared sequence
    sheet.cell(1, 2).value = "k-th item in the sequence"
    #And the fourth column is the position of the SNP in the sequence (k)
    sheet.cell(1, 3).value = "k"
    #"row" variable shows the current last row to put the new SNP there. So at the beginning, it is equal to 1 and its value is updated throghout the loop
    row = 2
    #The next loop adds the information of this SNP in to the Excel file based on the above mentioned definition. 
    for item in SNPs[key]:
        sheet.cell(row, 1).value = item[0]
        sheet.cell(row, 2).value = item[1]
        sheet.cell(row, 3).value = item[2]
        #Then we updated the "row" variable so that we are at the beginning of the next empty row to add information
        row = row + 1

#Finally we save the Excel workbook with the name of "Output.xlsx"
wb.save("Output.xlsx")
